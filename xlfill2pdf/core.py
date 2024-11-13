import os
import io
from pathlib import Path
import tempfile
from urllib.request import urlopen, quote
from typing import List, Optional, Union

import qrcode
import openpyxl
from PIL import Image as PILImage
from reportlab.platypus import Image
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

from .font import FontManager
from .qrcode import QRCodeGenerator


class ExcelProcessor:
    """Excel处理器，用于将Excel文件转换为PDF
    Excel processor for converting Excel files to PDF
    """

    # 图片相关常量 / Image related constants
    IMAGE_MAX_TOTAL_WIDTH = 350  # 最大总宽度 / Maximum total width
    IMAGE_TARGET_WIDTH = 80  # 目标宽度 / Target width
    IMAGE_HORIZONTAL_SPACING = 5  # 水平间距 / Horizontal spacing
    IMAGE_VERTICAL_SPACING = 5  # 垂直间距 / Vertical spacing
    IMAGE_MAX_PER_ROW = 3  # 每行最大图片数 / Maximum images per row

    # PDF相关常量 / PDF related constants
    PDF_MAX_IMAGE_WIDTH = 100  # PDF中图片最大宽度 / Maximum image width in PDF
    PDF_MAX_IMAGE_HEIGHT = 150  # PDF中图片最大高度 / Maximum image height in PDF

    def __init__(
        self,
        font_manager: FontManager,
        prefix: str = "{{",
        suffix: str = "}}",
        qrcode_suffix: str = ".qrcode",
        image_suffix: str = ".png",
        info_qrcode_suffix: str = ".info_qrcode",
        use_default_image_handlers: bool = True,
        use_default_qrcode_handlers: bool = True,
        use_default_info_qrcode_handlers: bool = True,
        watermark_text: Optional[str] = None,
        watermark_alpha: float = 0.1,
        watermark_angle: float = -45,
        watermark_color: tuple[int, int, int] = (0, 0, 0),
        qrcode_template: Optional[dict] = None,
    ):
        self.temporary_files = []
        self.font_manager = font_manager
        self.prefix = prefix
        self.suffix = suffix
        self.image_suffix = image_suffix
        self.qrcode_suffix = qrcode_suffix
        self.handlers = {}
        self.suffix_list = []
        self.watermark_text = watermark_text
        self.watermark_alpha = watermark_alpha
        self.watermark_angle = watermark_angle
        self.watermark_color = watermark_color
        self.qrcode_template = qrcode_template or {}
        self.info_qrcode_suffix = info_qrcode_suffix

        if use_default_qrcode_handlers:
            self.register_handler(self.qrcode_suffix, self._handle_qrcode)
        if use_default_image_handlers:
            self.register_handler(self.image_suffix, self._handle_image)
        if use_default_info_qrcode_handlers:
            self.register_handler(self.info_qrcode_suffix, self._handle_info_qrcode)
        self._register_font()

    def process_excel_to_pdf(self, excel_path: str, data_dict: dict) -> bytes:
        """处理Excel文件并转换为PDF
        Process Excel file and convert to PDF

        Args:
            excel_path: Excel文件路径 / Path to Excel file
            data_dict: 替换数据字典 / Dictionary containing replacement values

        Returns:
            bytes: 生成的PDF内容 / Generated PDF content
        """
        try:
            temp_excel_path = self._replace_placeholders(excel_path, data_dict)
            temp_pdf_path = self._excel_to_pdf(temp_excel_path)
            with open(temp_pdf_path, "rb") as pdf_file:
                return pdf_file.read()
        finally:
            for temp_file in self.temporary_files:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)

    def _register_font(self):
        """注册字体到reportlab系统
        Register font to reportlab system
        """
        try:
            font_path = self.font_manager.font_path
            pdfmetrics.registerFont(TTFont(self.font_manager.font_name, str(font_path)))
        except Exception as e:
            raise Exception(f"Font registration failed: {e}")

    def register_handler(self, suffix: str, handler_func):
        """注册自定义处理器
        Register custom handler

        Args:
            suffix: 处理器对应的后缀，如 ".qrcode", ".png" 等
                   Handler suffix, such as ".qrcode", ".png", etc.
            handler_func: 处理函数，接收 (cell, field_name, field_value, data_dict) 参数
                        Handler function that accepts (cell, field_name, field_value, data_dict) parameters
        """
        self.suffix_list.append(suffix)
        self.handlers[suffix] = handler_func

    def _handle_qrcode(self, cell, field_name, data_dict):
        """处理二维码的默认处理器"""
        qr_cord_img_path = self.generate_qr_code(data_dict.get(field_name))
        img = openpyxl.drawing.image.Image(qr_cord_img_path)
        img.width = 50
        img.height = 50
        cell.value = None
        cell.alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center"
        )

        column_letter = openpyxl.utils.get_column_letter(cell.column)
        anchor = f"{column_letter}{cell.row}"

        img.anchor = anchor
        return img, column_letter, cell.row

    def _handle_info_qrcode(self, cell, field_name, data_dict):
        """处理带信息的二维码
        Process QR code with additional information
        """
        # 创建二维码生成器实例
        qrc = QRCodeGenerator(
            font_manager=self.font_manager,
            qr_size=(50, 50),
            output_type="temp",
        )
        qr_path = qrc.create_info_qrcode(
            data_dict.get(field_name), self.qrcode_template
        )

        img = openpyxl.drawing.image.Image(qr_path)
        cell.value = None
        cell.alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center"
        )

        column_letter = openpyxl.utils.get_column_letter(cell.column)
        anchor = f"{column_letter}{cell.row}"

        img.anchor = anchor
        return img, column_letter, cell.row

    def _load_image_from_path_or_url(self, path: str):
        """从路径或URL加载图片
        Load image from path or URL

        Args:
            path: 图片路径或URL / Image path or URL

        Returns:
            PIL.Image: 加载的图片对象 / Loaded image object
        """
        try:
            if path.startswith("http"):
                # 处理 URL / Handle URL
                img_data = urlopen(path).read()
                return PILImage.open(io.BytesIO(img_data))
            else:
                # 处理本地文件路径 / Handle local file path
                return PILImage.open(path.strip())
        except Exception as e:
            print(f"Warning: Failed to load image {path}: {str(e)}")
            return None

    def _calc_row_width(self, image_count: int) -> int:
        """计算基于图片数量的行总宽度
        Calculate the total width of a row based on number of images

        Args:
            image_count: 图片数量 / Number of images

        Returns:
            int: 行总宽度 / Total row width
        """
        return (
            image_count * self.IMAGE_TARGET_WIDTH
            + (image_count - 1) * self.IMAGE_HORIZONTAL_SPACING
        )

    def _handle_image(self, cell, field_name, data_dict):
        """处理图片，支持多图片拼接
        Process images, supports multiple image concatenation

        图片路径可以用分号分隔，例如: "path1.png;path2.png;path3.png"
        Image paths can be separated by semicolons, e.g.: "path1.png;path2.png;path3.png"

        Args:
            cell: Excel单元格对象 / Excel cell object
            field_name: 字段名称 / Field name
            data_dict: 数据字典 / Data dictionary

        Returns:
            tuple: (图片对象, 列标识, 行号) / (Image object, column letter, row number)
        """
        image_paths = data_dict.get(field_name)
        if not image_paths:
            image_paths = []
        elif isinstance(image_paths, str):
            image_paths = [image_paths]

        # 调整所有图片大小并保持宽高比
        resized_images = []
        for path in image_paths:
            try:
                img = self._load_image_from_path_or_url(path.strip())
                if img:
                    aspect_ratio = img.height / img.width
                    target_height = int(self.IMAGE_TARGET_WIDTH * aspect_ratio)
                    resized_img = img.resize(
                        (self.IMAGE_TARGET_WIDTH, target_height),
                        PILImage.Resampling.LANCZOS,
                    )
                    resized_images.append(resized_img)
            except Exception as e:
                print(f"Warning: Failed to process image {path}: {str(e)}")
                continue

        if not resized_images:
            return None

        # 将图片分组到行
        rows = []
        current_row = []
        current_width = 0

        for img in resized_images:
            new_width = (
                current_width
                + self.IMAGE_TARGET_WIDTH
                + (len(current_row) > 0) * self.IMAGE_HORIZONTAL_SPACING
            )

            if (
                new_width > self.IMAGE_MAX_TOTAL_WIDTH
                or len(current_row) >= self.IMAGE_MAX_PER_ROW
            ):
                rows.append(current_row)
                current_row = [img]
                current_width = self.IMAGE_TARGET_WIDTH
            else:
                current_row.append(img)
                current_width = new_width

        # 添加最后一行
        if current_row:
            rows.append(current_row)

        # 计算每行的高度和总高度
        row_heights = []
        for row in rows:
            max_height = max(img.height for img in row)
            row_heights.append(max_height)

        total_height = sum(row_heights) + (len(rows) - 1) * self.IMAGE_VERTICAL_SPACING
        total_width = max(self._calc_row_width(len(row)) for row in rows)

        # 创建新的画布
        combined_image = PILImage.new("RGB", (total_width, total_height), "white")

        # 在画布上粘贴所有图片
        y_offset = 0
        for row, row_height in zip(rows, row_heights):
            x_offset = 0

            for img in row:
                # 在当前行内垂直居中
                y_pos = y_offset + (row_height - img.height) // 2
                combined_image.paste(img, (x_offset, y_pos))
                x_offset += self.IMAGE_TARGET_WIDTH + self.IMAGE_HORIZONTAL_SPACING

            y_offset += row_height + self.IMAGE_VERTICAL_SPACING

        # 保存合并后的图片
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
            combined_image.save(tmp.name)
            self.temporary_files.append(tmp.name)

            # 创建 openpyxl 图片对象
            excel_img = openpyxl.drawing.image.Image(tmp.name)

            # 设置合适的显示大小
            excel_img.width = total_width
            excel_img.height = total_height

            cell.value = None
            column_letter = openpyxl.utils.get_column_letter(cell.column)
            excel_img.anchor = f"{column_letter}{cell.row}"

            # 设置单元格高度
            cell.parent.row_dimensions[cell.row].height = total_height * 0.9

            return excel_img, column_letter, cell.row

    def generate_qr_code(self, data):
        """生成二维码图片
        Generate QR code image

        Args:
            data: 二维码数据内容 / QR code data content

        Returns:
            str: 生成的二维码图片路径 / Generated QR code image path
        """
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
            img.save(tmp.name)
            self.temporary_files.append(tmp.name)
            img_path = tmp.name
        return img_path

    def _replace_placeholders(self, excel_path: str, data_dict: dict):
        if excel_path.startswith("http"):
            try:
                encoded_url = quote(excel_path, safe=":/?=&")
                bytes_data = urlopen(encoded_url).read()
            except Exception as e:
                raise Exception(f"Download failed: {e}")
            wb = openpyxl.load_workbook(io.BytesIO(bytes_data))
        else:
            wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active

        for row in sheet.iter_rows():
            for cell in row:
                if (
                    cell.value
                    and isinstance(cell.value, str)
                    and cell.value.startswith(self.prefix)
                    and cell.value.endswith(self.suffix)
                ):
                    placeholder = cell.value[len(self.prefix) : -len(self.suffix)]
                    if placeholder in data_dict:
                        cell.value = data_dict[placeholder]
                    else:
                        for handle_suffix in self.suffix_list:
                            flag = False
                            for current_suffix in handle_suffix.split(","):
                                if placeholder.endswith(current_suffix):
                                    field_name = placeholder[: -len(current_suffix)]
                                    handler_func_obj = self.handlers.get(handle_suffix)
                                    # 调用对应的处理器
                                    result = handler_func_obj(
                                        cell, field_name, data_dict
                                    )
                                    if result:
                                        if (
                                            isinstance(result, tuple)
                                            and len(result) == 3
                                        ):  # Image handler result
                                            img, column_letter, row_num = result
                                            sheet.add_image(img)
                                            sheet.column_dimensions[
                                                column_letter
                                            ].width = (img.width / 9)
                                            sheet.row_dimensions[row_num].height = (
                                                img.height * 0.9
                                            )
                                        else:
                                            cell.value = str(result)
                                    flag = True
                                    break
                            if flag:
                                break

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            self.temporary_files.append(tmp.name)
            tmp_path = tmp.name
        return tmp_path

    def _add_watermark(self, canvas_obj, pagesize):
        """添加水印到PDF页面
        Add watermark to PDF page

        Args:
            canvas_obj: PDF画布对象 / PDF canvas object
            pagesize: 页面尺 / Page size
        """
        if not self.watermark_text:
            return

        canvas_obj.saveState()

        # 设置水印文字属性 / Set watermark text properties
        canvas_obj.setFont(self.font_manager.font_name, 60)
        r, g, b = self.watermark_color
        canvas_obj.setFillColorRGB(r, g, b, alpha=self.watermark_alpha)

        # 计算水印位置和旋转 / Calculate watermark position and rotation
        page_width, page_height = pagesize

        # 创建水印网格 / Create watermark grid
        text_width = canvas_obj.stringWidth(
            self.watermark_text, self.font_manager.font_name, 60
        )
        text_height = 60  # 假设高度为字体大小 / Assume height equals font size

        # 计算水印间距 / Calculate watermark spacing
        x_spacing = text_width * 2
        y_spacing = text_height * 2

        # 在页面上绘制水印网格 / Draw watermark grid on page
        for y in range(0, int(page_height * 1.5), int(y_spacing)):
            for x in range(0, int(page_width * 1.5), int(x_spacing)):
                canvas_obj.saveState()
                canvas_obj.translate(x, y)
                canvas_obj.rotate(self.watermark_angle)
                canvas_obj.drawString(0, 0, self.watermark_text)
                canvas_obj.restoreState()

        canvas_obj.restoreState()

    def _excel_to_pdf(self, excel_path: str):
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        tmp_pdf_fd, tmp_pdf_path = tempfile.mkstemp(suffix=".pdf")
        self.temporary_files.append(tmp_pdf_path)
        os.close(tmp_pdf_fd)

        margin = 0.3 * inch
        doc = SimpleDocTemplate(
            tmp_pdf_path,
            pagesize=landscape(letter),
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
        )
        data = []
        merged_cells = sheet.merged_cells.ranges
        for row_index, row in enumerate(sheet.iter_rows(), start=1):
            row_data = []
            for col_index, cell in enumerate(row, start=1):
                value = cell.value if cell.value is not None else ""
                for merged_range in merged_cells:
                    if cell.coordinate in merged_range:
                        if cell.coordinate == merged_range.start_cell.coordinate:
                            value = (
                                merged_range.start_cell.value
                                if merged_range.start_cell.value is not None
                                else ""
                            )
                        else:
                            value = ""
                        break

                # 处理图片
                image = None
                for img in sheet._images:
                    if (
                        img.anchor._from.row == row_index - 1
                        and img.anchor._from.col == col_index - 1
                    ):
                        image = img
                        break

                if image:
                    try:
                        img_data = image.ref
                        if hasattr(img_data, "getvalue"):
                            img_bytes = img_data.getvalue()
                        else:
                            img_bytes = img_data

                        pil_img = PILImage.open(io.BytesIO(img_bytes))
                        img_width, img_height = pil_img.size
                        aspect = img_height / float(img_width)

                        # 计算缩放后的尺寸
                        img_width = min(img_width, self.PDF_MAX_IMAGE_WIDTH)
                        img_height = img_width * aspect

                        # 如果高度超过限制，从度反向计算宽度
                        if img_height > self.PDF_MAX_IMAGE_HEIGHT:
                            img_height = self.PDF_MAX_IMAGE_HEIGHT
                            img_width = img_height / aspect

                        with tempfile.NamedTemporaryFile(
                            delete=False, suffix=".png"
                        ) as temp_img_file:
                            self.temporary_files.append(temp_img_file.name)
                            # 保存调整后的图片
                            pil_img = pil_img.resize(
                                (int(img_width), int(img_height)),
                                PILImage.Resampling.LANCZOS,
                            )
                            pil_img.save(temp_img_file.name)
                            value = Image(
                                temp_img_file.name, width=img_width, height=img_height
                            )
                    except Exception as e:
                        value = "write failed"

                row_data.append(value)
            if any(cell != "" for cell in row_data):
                data.append(row_data)

        if not data:
            data = [[""]]
        style = ParagraphStyle(
            "Normal",
            fontName=self.font_manager.font_name,  # 使用字体管理器的字体
            fontSize=8,
            leading=10,
            alignment=1,
        )
        data = [
            [
                Paragraph(str(cell), style) if isinstance(cell, str) else cell
                for cell in row
            ]
            for row in data
        ]

        # 创建一个字典来存储每行的最大高度
        row_heights = {}

        # 首先获取Excel中设置的原始行高
        for i in range(1, len(data) + 1):
            if i in sheet.row_dimensions:
                row_heights[i] = sheet.row_dimensions[i].height

        # 检查每个单元格是否包含图片，并更新行高
        for row_index, row in enumerate(data, start=1):
            for col_index, cell in enumerate(row):
                if isinstance(cell, Image):
                    # 如果单元格包含图片，检查并更新该行的高度
                    image_height = cell.drawHeight
                    current_height = row_heights.get(row_index, 0)
                    # 将图片高度转换为与Excel兼容的单位
                    excel_image_height = image_height * 1.2  # 添加一些额外空间
                    if excel_image_height > current_height:
                        row_heights[row_index] = excel_image_height

        # 使用更新后的行高创建Table
        table = Table(
            data, rowHeights=[row_heights.get(i, None) for i in range(1, len(data) + 1)]
        )
        table_style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTNAME", (0, 0), (-1, -1), self.font_manager.font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ]
        )
        for merged_range in merged_cells:
            table_style.add(
                "SPAN",
                (merged_range.min_col - 1, merged_range.min_row - 1),
                (merged_range.max_col - 1, merged_range.max_row - 1),
            )
        table.setStyle(table_style)
        elements = [table]

        # 修改 WatermarkCanvas 类的实现
        class WatermarkCanvas(canvas.Canvas):
            """水印画布类，用于在PDF页面上添加水印
            Watermark canvas class for adding watermarks to PDF pages
            """

            def __init__(self, filename, processor=None, **kwargs):
                super().__init__(filename, **kwargs)
                self.processor = processor

            def showPage(self):
                """显示页面并添加水印
                Show page and add watermark
                """
                if self.processor and self.processor.watermark_text:
                    self.processor._add_watermark(self, landscape(letter))
                super().showPage()

        # 修改 doc.build 调用，传递所有参数
        doc.build(
            elements,
            canvasmaker=lambda filename, **kwargs: WatermarkCanvas(
                filename, processor=self, **kwargs
            ),
        )
        return tmp_pdf_path
